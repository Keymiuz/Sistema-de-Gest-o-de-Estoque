import sqlite3
import pandas as pd
from database import DATABASE_PATH
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def conectar_bd():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    return conn, cursor

def calcular_total_vendas(data_inicio=None, data_fim=None):
    conn, cursor = conectar_bd()
    total = 0.0
    try:
        query = "SELECT SUM(valor_total) as total_vendas FROM faturamento"
        params = []
        if data_inicio and data_fim:
            query += " WHERE date(data) BETWEEN date(?) AND date(?)" 
            params.extend([data_inicio, data_fim])
        
        cursor.execute(query, params)
        resultado = cursor.fetchone()
        if resultado and resultado['total_vendas'] is not None:
            total = resultado['total_vendas']
    except Exception as e:
        print(f"Erro ao calcular total de vendas: {e}")
    finally:
        conn.close()
    return total

def calcular_total_compras(data_inicio=None, data_fim=None): 
    conn, cursor = conectar_bd()
    total = 0.0
    try:
        query = "SELECT SUM(quantidade * preco_custo_unitario) as total_compras FROM compras_historico"
        params = []
        if data_inicio and data_fim:
            query += " WHERE date(data_compra) BETWEEN date(?) AND date(?)"
            params.extend([data_inicio, data_fim])

        cursor.execute(query, params)
        resultado = cursor.fetchone()
        if resultado and resultado['total_compras'] is not None:
            total = resultado['total_compras']
    except Exception as e:
        print(f"Erro ao calcular total de compras: {e}")
    finally:
        conn.close()
    return total

def calcular_total_despesas(data_inicio=None, data_fim=None):
    conn, cursor = conectar_bd()
    total = 0.0
    try:
        query = "SELECT SUM(valor) as total_despesas FROM despesas"
        params = []
        if data_inicio and data_fim:
            query += " WHERE date(data) BETWEEN date(?) AND date(?)"
            params.extend([data_inicio, data_fim])

        cursor.execute(query, params)
        resultado = cursor.fetchone()
        if resultado and resultado['total_despesas'] is not None:
            total = resultado['total_despesas']
    except Exception as e:
        print(f"Erro ao calcular total de despesas: {e}")
    finally:
        conn.close()
    return total

def gerar_relatorio_financeiro_consolidado(data_inicio=None, data_fim=None, detalhado=False):
    # Dados consolidados
    vendas = calcular_total_vendas(data_inicio, data_fim)
    compras = calcular_total_compras(data_inicio, data_fim)
    despesas = calcular_total_despesas(data_inicio, data_fim)
    saldo = vendas - compras - despesas
    
    if not detalhado:
        return {
            "total_vendas": vendas,
            "total_compras_custo": compras,
            "total_despesas": despesas,
            "saldo_empresarial": saldo
        }
    else:
        # Se for detalhado, retornar também os dados completos para exportação
        conn, cursor = conectar_bd()
        try:
            # Buscar vendas detalhadas
            query_vendas = "SELECT data, valor_total FROM faturamento"
            params = []
            if data_inicio and data_fim:
                query_vendas += " WHERE date(data) BETWEEN date(?) AND date(?)"
                params.extend([data_inicio, data_fim])
            cursor.execute(query_vendas, params)
            vendas_detalhadas = [dict(row) for row in cursor.fetchall()]
            
            # Buscar compras detalhadas
            query_compras = """
                SELECT data_compra as data, 
                       SUM(quantidade * preco_custo_unitario) as valor_total
                FROM compras_historico
            """
            if data_inicio and data_fim:
                query_compras += " WHERE date(data_compra) BETWEEN date(?) AND date(?)"
            query_compras += " GROUP BY data_compra"
            cursor.execute(query_compras, params if data_inicio and data_fim else [])
            compras_detalhadas = [dict(row) for row in cursor.fetchall()]
            
            # Buscar despesas detalhadas
            query_despesas = "SELECT data, valor as valor_total, descricao FROM despesas"
            if data_inicio and data_fim:
                query_despesas += " WHERE date(data) BETWEEN date(?) AND date(?)"
            cursor.execute(query_despesas, params if data_inicio and data_fim else [])
            despesas_detalhadas = [dict(row) for row in cursor.fetchall()]
            
            return {
                "resumo": {
                    "total_vendas": vendas,
                    "total_compras_custo": compras,
                    "total_despesas": despesas,
                    "saldo_empresarial": saldo
                },
                "detalhes": {
                    "vendas": vendas_detalhadas,
                    "compras": compras_detalhadas,
                    "despesas": despesas_detalhadas
                },
                "periodo": {
                    "data_inicio": data_inicio if data_inicio else "-",
                    "data_fim": data_fim if data_fim else "-"
                }
            }
            
        except Exception as e:
            print(f"Erro ao buscar dados detalhados: {e}")
            return None
        finally:
            conn.close()

def exportar_para_excel(data_inicio=None, data_fim=None, diretorio_saida=None):
    """
    Exporta o relatório financeiro para um arquivo Excel com fórmulas de soma e formatação aprimorada.
    
    Args:
        data_inicio (str, optional): Data de início no formato 'YYYY-MM-DD'
        data_fim (str, optional): Data de fim no formato 'YYYY-MM-DD'
        diretorio_saida (str, optional): Diretório para salvar o arquivo. Se não informado, usa o diretório do usuário.
    
    Returns:
        str: Caminho completo do arquivo gerado ou None em caso de erro
    """
    try:
        # Obter dados detalhados
        dados = gerar_relatorio_financeiro_consolidado(data_inicio, data_fim, detalhado=True)
        if not dados:
            return None
        
        # Criar nome do arquivo com data e hora
        data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_financeiro_{data_hora}.xlsx"
        
        # Definir diretório de saída
        if not diretorio_saida:
            diretorio_saida = os.path.expanduser("~")
        
        caminho_arquivo = os.path.join(diretorio_saida, nome_arquivo)
        
        # Estilos comuns
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=11, bold=True)
        currency_format = 'R$ #,##0.00' # Formato de moeda brasileiro
        date_format = 'dd/mm/yyyy'

        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            # ===== PLANILHA DE RESUMO =====
            # Os valores aqui serão fórmulas que buscam os totais das outras planilhas
            resumo_data = {
                'Descrição': [
                    "TOTAL DE VENDAS (RECEITA)",
                    "TOTAL GASTO EM COMPRAS",
                    "TOTAL DE DESPESAS",
                    "",
                    "SALDO EMPRESARIAL"
                ],
                'Valor': [
                    "=SUM(Vendas!B:B)",  # Soma a coluna B da planilha Vendas
                    "=SUM(Compras!B:B)", # Soma a coluna B da planilha Compras
                    "=SUM(Despesas!C:C)",# Soma a coluna C da planilha Despesas (considerando 'valor_total')
                    "",
                    "=B2-B3-B4"          # Saldo: Vendas - Compras - Despesas
                ]
            }
            resumo_df = pd.DataFrame(resumo_data)
            resumo_df.to_excel(writer, sheet_name='Resumo', index=False, header=False, startrow=2)
            
            ws_resumo = writer.sheets['Resumo']
            ws_resumo.column_dimensions['A'].width = 30
            ws_resumo.column_dimensions['B'].width = 20
            
            periodo_str = f"Período: {dados['periodo']['data_inicio']} a {dados['periodo']['data_fim']}" if dados['periodo']['data_inicio'] != "-" else "Período: Todos os dados"
            ws_resumo.cell(row=1, column=1, value="RELATÓRIO FINANCEIRO CONSOLIDADO").font = title_font
            ws_resumo.cell(row=2, column=1, value=periodo_str).font = Font(name='Arial', size=10, italic=True)

            for row_idx in range(3, 8): # Linhas 3 a 7 (dados do resumo, considerando o startrow=2)
                cell_desc = ws_resumo.cell(row=row_idx, column=1)
                cell_val = ws_resumo.cell(row=row_idx, column=2)
                cell_desc.border = thin_border
                cell_val.border = thin_border
                cell_val.number_format = currency_format
                if row_idx == 7: # Saldo
                    cell_desc.font = header_font
                    cell_val.font = header_font
            
            # Proteger a planilha de Resumo
            ws_resumo.protection.sheet = True

            # ===== PLANILHAS DETALHADAS =====
            def format_detailed_sheet(sheet_name, df_data, value_column_letter='B'):
                if not df_data.empty:
                    df_data.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]
                    
                    # Formatar cabeçalho
                    for col_idx, column_title in enumerate(df_data.columns, 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Ajustar largura das colunas e formatar dados
                    for col_idx, column in enumerate(df_data.columns, 1):
                        max_len = max(df_data[column].astype(str).map(len).max(), len(column)) + 2
                        ws.column_dimensions[chr(64 + col_idx)].width = max_len
                        
                        for row_idx in range(2, len(df_data) + 2):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border
                            if df_data[column].dtype == 'datetime64[ns]' or column.lower() == 'data':
                                cell.number_format = date_format
                            elif column.lower().startswith('valor') or column.lower().startswith('preco') or column.lower().startswith('subtotal'):
                                cell.number_format = currency_format
                    
                    # Adicionar linha de TOTAL
                    total_row_idx = len(df_data) + 2
                    total_label_cell = ws.cell(row=total_row_idx, column=df_data.columns.get_loc(df_data.columns[df_data.columns.str.lower().str.contains('valor|preco|subtotal')][0])) # Coluna antes do valor
                    total_label_cell.value = "TOTAL:"
                    total_label_cell.font = header_font
                    total_label_cell.border = thin_border
                    total_label_cell.alignment = Alignment(horizontal='right')

                    total_value_cell = ws.cell(row=total_row_idx, column=df_data.columns.get_loc(df_data.columns[df_data.columns.str.lower().str.contains('valor|preco|subtotal')][0]) + 1)
                    total_value_cell.value = f"=SUM({value_column_letter}2:{value_column_letter}{total_row_idx-1})"
                    total_value_cell.font = header_font
                    total_value_cell.border = thin_border
                    total_value_cell.number_format = currency_format
                    ws.freeze_panes = 'A2'
                    
                    # Proteger a planilha detalhada
                    ws.protection.sheet = True

            # Vendas
            if dados['detalhes']['vendas']:
                vendas_df = pd.DataFrame(dados['detalhes']['vendas'])
                if not vendas_df.empty:
                    vendas_df = vendas_df.rename(columns={'data': 'Data', 'valor_total': 'Valor Venda (R$)'})
                    vendas_df['Data'] = pd.to_datetime(vendas_df['Data'])
                    format_detailed_sheet('Vendas', vendas_df, 'B')
            else:
                 pd.DataFrame([{'Data': None, 'Valor Venda (R$)': None}]).to_excel(writer, sheet_name='Vendas', index=False) # Planilha vazia para a fórmula do resumo funcionar

            # Compras
            if dados['detalhes']['compras']:
                compras_df = pd.DataFrame(dados['detalhes']['compras'])
                if not compras_df.empty:
                    compras_df = compras_df.rename(columns={'data': 'Data Compra', 'valor_total': 'Valor Compra (R$)'})
                    compras_df['Data Compra'] = pd.to_datetime(compras_df['Data Compra'])
                    format_detailed_sheet('Compras', compras_df, 'B')
            else:
                pd.DataFrame([{'Data Compra': None, 'Valor Compra (R$)': None}]).to_excel(writer, sheet_name='Compras', index=False)

            # Despesas
            if dados['detalhes']['despesas']:
                despesas_df = pd.DataFrame(dados['detalhes']['despesas'])
                if not despesas_df.empty:
                    despesas_df = despesas_df.rename(columns={'data': 'Data', 'valor_total': 'Valor Despesa (R$)', 'descricao': 'Descrição'})
                    despesas_df['Data'] = pd.to_datetime(despesas_df['Data'])
                    despesas_df = despesas_df[['Data', 'Descrição', 'Valor Despesa (R$)']]
                    format_detailed_sheet('Despesas', despesas_df, 'C') # Valor está na coluna C
            else:
                pd.DataFrame([{'Data': None, 'Descrição': None, 'Valor Despesa (R$)': None}]).to_excel(writer, sheet_name='Despesas', index=False)
                # Proteger planilha vazia também, se criada
                if 'Despesas' in writer.sheets:
                    writer.sheets['Despesas'].protection.sheet = True

        return caminho_arquivo
        
    except Exception as e:
        print(f"Erro ao exportar para Excel: {e}")
        # Considerar logar o traceback completo para depuração
        import traceback
        traceback.print_exc()
        return None

if __name__ == '__main__':
    relatorio = gerar_relatorio_financeiro_consolidado() 
    print("--- Relatório Financeiro Consolidado (Teste) ---")
    print(f"Total de Vendas: R$ {relatorio['total_vendas']:.2f}")
    print(f"Total Gasto em Compras: R$ {relatorio['total_compras_custo']:.2f}")
    print(f"Total de Despesas: R$ {relatorio['total_despesas']:.2f}")
    print(f"Saldo Empresarial: R$ {relatorio['saldo_empresarial']:.2f}")